"""Reading construction records out of spreadsheets.

Construction commercial documents are tables far more often than prose: a bill of quantities, a
measurement sheet, and a running account bill are all row-per-item grids. So extraction here is
column mapping rather than pattern matching, and the evidence a fact carries is a cell reference
rather than a character span — ``BOQ!D6`` is the spreadsheet equivalent of "page 6, characters
1204-1218", and it is what makes the value checkable.

Header matching is by known aliases, deliberately not by position. Real records reorder columns,
and a positional reader silently swaps quantity for rate the first time someone inserts a column —
which is the class of error that produces a plausible number from the wrong cell.

**A row is taken only if it yields an item identifier and at least one usable value.** Banner rows,
titles, blank spacers and totals all sit in the same sheet as the data, and reading them as items
would invent work that nobody contracted for.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Final

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from aedifex.domain.evidence import FactKind
from aedifex.errors import ExtractionError

__all__ = [
    "FIELD_CLAIMED_RATE",
    "FIELD_CONTRACTED_QUANTITY",
    "FIELD_CONTRACT_RATE",
    "FIELD_CUMULATIVE_CLAIM_QUANTITY",
    "FIELD_CURRENT_CLAIM_QUANTITY",
    "FIELD_ITEM_DESCRIPTION",
    "FIELD_ITEM_IDENTIFIER",
    "FIELD_MEASURED_QUANTITY",
    "FIELD_PREVIOUS_CERTIFIED_QUANTITY",
    "FIELD_PROJECT_REFERENCE",
    "MAX_WINDOW_COLUMNS",
    "MAX_WINDOW_ROWS",
    "SheetCell",
    "SheetFact",
    "SheetRow",
    "SheetWindow",
    "SheetWindowCell",
    "SheetWindowRow",
    "cell_reference",
    "read_construction_sheet",
    "read_sheet_window",
]

FIELD_ITEM_IDENTIFIER: Final[str] = "item_identifier"
FIELD_ITEM_DESCRIPTION: Final[str] = "item_description"
FIELD_CONTRACTED_QUANTITY: Final[str] = "contracted_quantity"
FIELD_CONTRACT_RATE: Final[str] = "contract_rate"
FIELD_MEASURED_QUANTITY: Final[str] = "measured_quantity"
FIELD_PREVIOUS_CERTIFIED_QUANTITY: Final[str] = "previous_certified_quantity"
FIELD_CURRENT_CLAIM_QUANTITY: Final[str] = "current_claim_quantity"
FIELD_CUMULATIVE_CLAIM_QUANTITY: Final[str] = "cumulative_claim_quantity"
FIELD_CLAIMED_RATE: Final[str] = "claimed_rate"
FIELD_PROJECT_REFERENCE: Final[str] = "project_reference"

CURRENCY_INR: Final[str] = "INR"

# Header text -> fact type. Matched against a normalised header, so "Cumulative Claim Quantity",
# "cumulative claim qty" and "CUMULATIVE  CLAIMED   QUANTITY" all land in one place.
#
# The rate columns are ambiguous by nature: a bill of quantities and a running bill both label the
# column "Rate", and they mean different things -- what was agreed versus what is being claimed. So
# the rate alias resolves per document type rather than globally, in _rate_field_for.
_HEADER_ALIASES: Final[dict[str, str]] = {
    "item no": FIELD_ITEM_IDENTIFIER,
    "item number": FIELD_ITEM_IDENTIFIER,
    "item": FIELD_ITEM_IDENTIFIER,
    "sl no": FIELD_ITEM_IDENTIFIER,
    "description": FIELD_ITEM_DESCRIPTION,
    "particulars": FIELD_ITEM_DESCRIPTION,
    "quantity": FIELD_CONTRACTED_QUANTITY,
    "qty": FIELD_CONTRACTED_QUANTITY,
    "contracted quantity": FIELD_CONTRACTED_QUANTITY,
    "measured quantity": FIELD_MEASURED_QUANTITY,
    "measured qty": FIELD_MEASURED_QUANTITY,
    "previous certified quantity": FIELD_PREVIOUS_CERTIFIED_QUANTITY,
    "previous certified qty": FIELD_PREVIOUS_CERTIFIED_QUANTITY,
    "previous quantity": FIELD_PREVIOUS_CERTIFIED_QUANTITY,
    "current claim quantity": FIELD_CURRENT_CLAIM_QUANTITY,
    "current quantity": FIELD_CURRENT_CLAIM_QUANTITY,
    "cumulative claim quantity": FIELD_CUMULATIVE_CLAIM_QUANTITY,
    "cumulative quantity": FIELD_CUMULATIVE_CLAIM_QUANTITY,
    "cumulative claimed quantity": FIELD_CUMULATIVE_CLAIM_QUANTITY,
}

_UNIT_HEADERS: Final[frozenset[str]] = frozenset({"unit", "units", "uom"})
_RATE_HEADERS: Final[frozenset[str]] = frozenset({"rate", "rate (rs)", "unit rate"})

# Quantities and money, so a fact knows what it is without consulting its name.
_QUANTITY_FIELDS: Final[frozenset[str]] = frozenset(
    {
        FIELD_CONTRACTED_QUANTITY,
        FIELD_MEASURED_QUANTITY,
        FIELD_PREVIOUS_CERTIFIED_QUANTITY,
        FIELD_CURRENT_CLAIM_QUANTITY,
        FIELD_CUMULATIVE_CLAIM_QUANTITY,
    }
)
_MONEY_FIELDS: Final[frozenset[str]] = frozenset({FIELD_CONTRACT_RATE, FIELD_CLAIMED_RATE})

# An item identifier looks like 4.7.2 or 12 or A-3. Anything else in that column is a banner, a
# heading or a total, and must not become a work item.
_ITEM_IDENTIFIER: Final[re.Pattern[str]] = re.compile(r"^[A-Za-z]?[0-9]+(?:[.\-/][0-9A-Za-z]+)*$")

_PROJECT_LINE: Final[re.Pattern[str]] = re.compile(
    r"project\s*(?:reference|ref|no|number)?\s*[:\-]\s*(?P<value>[A-Za-z0-9][A-Za-z0-9\-_/]*)",
    re.IGNORECASE,
)

MAX_ROWS: Final[int] = 5_000
MAX_COLUMNS: Final[int] = 64


def cell_reference(sheet: str, row: int, column: int) -> str:
    """``BOQ!D6`` — what a reviewer types into the go-to box to see it.

    A module-level function as well as a property, because the API has to build the same string from
    three stored columns and the format must be defined once. A viewer showing ``BOQ!D6`` while the
    extractor recorded ``BOQ.D6`` would be two answers to "where is this?".
    """
    return f"{sheet}!{get_column_letter(column)}{row}"


@dataclass(frozen=True, slots=True)
class SheetCell:
    """Where a value came from, in the terms a spreadsheet uses."""

    sheet: str
    row: int
    column: int

    @property
    def reference(self) -> str:
        """``BOQ!D6`` — what a reviewer types into the go-to box to see it."""
        return cell_reference(self.sheet, self.row, self.column)


@dataclass(frozen=True, slots=True)
class SheetFact:
    """One value read from one cell."""

    fact_type: str
    kind: FactKind
    literal: str
    value: Decimal | None
    unit: str | None
    currency: str | None
    cell: SheetCell


@dataclass(frozen=True, slots=True)
class SheetRow:
    """One work item's row: its identifier and every value found alongside it."""

    item_identifier: str
    row_number: int
    facts: tuple[SheetFact, ...]

    def fact(self, fact_type: str) -> SheetFact | None:
        for candidate in self.facts:
            if candidate.fact_type == fact_type:
                return candidate
        return None


@dataclass(frozen=True, slots=True)
class ConstructionSheet:
    """What one spreadsheet yielded."""

    sheet_name: str
    project_reference: SheetFact | None
    rows: tuple[SheetRow, ...]
    unsupported: tuple[str, ...]


def _normalise_header(value: object) -> str:
    return " ".join(str(value).strip().lower().replace("(rs)", "").split()) if value else ""


def _to_decimal(value: object) -> Decimal | None:
    """Convert a cell value to an exact Decimal, or refuse.

    Goes via ``str`` for floats on purpose: openpyxl returns numeric cells as Python floats, and
    ``Decimal(0.1)`` is not ``Decimal("0.1")``. Constructing from the repr keeps the number the
    spreadsheet displays rather than the binary approximation behind it.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _rate_field_for(document_type: str) -> str:
    """Which fact a bare ``Rate`` column means, given what kind of document it is.

    A bill of quantities states the contracted rate; a running bill states the rate being claimed.
    The column is labelled identically in both, so the document type is the only thing that
    disambiguates it — and reading a claimed rate as a contracted one would make a rate variance
    invisible.
    """
    return FIELD_CLAIMED_RATE if document_type == "running_bill" else FIELD_CONTRACT_RATE


MAX_WINDOW_ROWS: Final[int] = 61
"""Rows a single window may return: 30 either side of the target, plus the target.

Bounded because the caller is an HTTP request and the sheet is untrusted input. Wide enough that a
reviewer sees the header of a bill along with the row in question when they are near each other, and
narrow enough that no response is a table nobody reads.
"""

MAX_WINDOW_COLUMNS: Final[int] = 26
_MAX_CELL_CHARS: Final[int] = 200


@dataclass(frozen=True, slots=True)
class SheetWindowCell:
    """One cell of a window: where it is, and what it says."""

    column: int
    letter: str
    reference: str
    value: str


@dataclass(frozen=True, slots=True)
class SheetWindowRow:
    number: int
    cells: tuple[SheetWindowCell, ...]


@dataclass(frozen=True, slots=True)
class SheetWindow:
    """A readable region of one sheet, for showing a reviewer where a value came from.

    **Not an extraction.** Nothing here becomes a fact, nothing is normalised, and no column is
    interpreted — this is the spreadsheet equivalent of opening a PDF at page 6. Values are rendered
    as the text of what ``openpyxl`` read, using the same library and the same ``data_only`` setting
    the extractor used, so the window cannot disagree with the facts about what a cell contains. A
    second reader in the browser could, which is the reason this lives here rather than there.
    """

    sheet: str
    sheets: tuple[str, ...]
    first_row: int
    rows: tuple[SheetWindowRow, ...]
    total_rows: int
    truncated: bool
    """Whether the sheet has rows outside the window. Never a silent crop."""


def read_sheet_window(
    data: bytes,
    *,
    sheet: str | None = None,
    row: int = 1,
    radius: int = 12,
    columns: int = MAX_WINDOW_COLUMNS,
) -> SheetWindow:
    """Read a bounded region of a spreadsheet around ``row``, for display only.

    The authoritative artifact is still the workbook, and it stays downloadable. This exists because
    the alternative was worse in both directions: leaving spreadsheet evidence unviewable made the
    *strongest* evidence the project holds the hardest to check, and rendering it with a second
    parser in the browser would let the viewer and the extractor disagree about what cell F43 says.

    Args:
        sheet: Sheet by name. Defaults to the first, which is what the extractor reads.
        row: The row to centre on — the one a fact cites.
        radius: Rows either side. Clamped so a window is always bounded.
        columns: Columns from A. Clamped to ``MAX_WINDOW_COLUMNS``.

    Raises:
        ExtractionError: if the workbook cannot be opened, has no worksheets, or has no such sheet.
    """
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as error:
        raise ExtractionError(
            f"spreadsheet could not be opened: {type(error).__name__}: {error}"
        ) from error

    try:
        names = tuple(workbook.sheetnames)
        if not names:
            raise ExtractionError("spreadsheet contains no worksheets")
        if sheet is not None and sheet not in names:
            raise ExtractionError(f"no sheet named {sheet!r}; this workbook has {', '.join(names)}")
        worksheet = workbook[sheet] if sheet is not None else workbook[names[0]]

        width = max(1, min(columns, MAX_WINDOW_COLUMNS))
        span = max(0, min(radius, MAX_WINDOW_ROWS // 2))
        first = max(1, row - span)
        last = first + min(MAX_WINDOW_ROWS, span * 2 + 1) - 1

        window: list[SheetWindowRow] = []
        seen = 0
        for index, values in enumerate(worksheet.iter_rows(values_only=True), start=1):
            seen = index
            if index > MAX_ROWS:
                break
            if index < first or index > last:
                continue
            cells = tuple(
                SheetWindowCell(
                    column=column,
                    letter=get_column_letter(column),
                    reference=cell_reference(worksheet.title, index, column),
                    value=_render(value),
                )
                for column, value in enumerate(values[:width], start=1)
            )
            window.append(SheetWindowRow(number=index, cells=cells))
    finally:
        workbook.close()

    return SheetWindow(
        sheet=worksheet.title,
        sheets=names,
        first_row=first,
        rows=tuple(window),
        total_rows=seen,
        truncated=seen > last or first > 1,
    )


def _render(value: object) -> str:
    """A cell as text, for display. Bounded, and never reformatted into something else.

    ``str`` of what openpyxl read, not a locale-formatted number: a viewer showing ``85,39,81,318``
    where the workbook holds ``853981318.41`` would be presenting our formatting as the document's
    content, which is the same mistake as re-rendering a PDF.
    """
    if value is None:
        return ""
    text = str(value)
    return text if len(text) <= _MAX_CELL_CHARS else f"{text[:_MAX_CELL_CHARS]}…"


def read_construction_sheet(data: bytes, *, document_type: str) -> ConstructionSheet:
    """Read one construction spreadsheet into rows of facts.

    Args:
        data: The raw XLSX bytes, as stored.
        document_type: ``bill_of_quantities``, ``measurement_book`` or ``running_bill``. Decides
            what an unlabelled ``Rate`` column means.

    Raises:
        ExtractionError: if the workbook cannot be opened or has no readable sheet.
    """
    try:
        # read_only bounds memory on a large sheet; data_only takes computed values rather than
        # formula text, because a claim of "=D6*E6" is not a number anyone can reconcile.
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as error:
        raise ExtractionError(
            f"spreadsheet could not be opened: {type(error).__name__}: {error}"
        ) from error

    try:
        sheet = workbook.worksheets[0] if workbook.worksheets else None
        if sheet is None:
            raise ExtractionError("spreadsheet contains no worksheets")
        sheet_name = sheet.title

        grid: list[list[object]] = []
        for index, row in enumerate(sheet.iter_rows(values_only=True)):
            if index >= MAX_ROWS:
                break
            grid.append(list(row[:MAX_COLUMNS]))
    finally:
        workbook.close()

    project_reference = _find_project_reference(grid, sheet_name)
    header_row, columns = _find_header(grid, document_type)
    if header_row is None:
        return ConstructionSheet(
            sheet_name=sheet_name,
            project_reference=project_reference,
            rows=(),
            unsupported=(
                "no header row recognised: none of the rows carried an item-number column "
                "alongside a quantity or rate column",
            ),
        )

    rows = tuple(_read_rows(grid, header_row, columns, sheet_name))
    unsupported: tuple[str, ...] = ()
    if not rows:
        unsupported = ("a header row was found but no row below it yielded an item identifier",)
    return ConstructionSheet(
        sheet_name=sheet_name,
        project_reference=project_reference,
        rows=rows,
        unsupported=unsupported,
    )


def _find_project_reference(grid: list[list[object]], sheet_name: str) -> SheetFact | None:
    """The project reference these records belong to, from a ``Project: X`` line."""
    for row_index, row in enumerate(grid, start=1):
        for column_index, cell in enumerate(row, start=1):
            if cell is None:
                continue
            match = _PROJECT_LINE.search(str(cell))
            if match is None:
                continue
            return SheetFact(
                fact_type=FIELD_PROJECT_REFERENCE,
                kind=FactKind.IDENTIFIER,
                literal=match.group("value"),
                value=None,
                unit=None,
                currency=None,
                cell=SheetCell(sheet=sheet_name, row=row_index, column=column_index),
            )
    return None


def _find_header(grid: list[list[object]], document_type: str) -> tuple[int | None, dict[int, str]]:
    """Locate the header row and map each column index to a fact type.

    A row qualifies only if it names an item column *and* at least one value column. Construction
    sheets open with banners and titles, and a looser test matches the first row holding any text.
    """
    rate_field = _rate_field_for(document_type)
    for row_index, row in enumerate(grid):
        columns: dict[int, str] = {}
        for column_index, cell in enumerate(row, start=1):
            header = _normalise_header(cell)
            if not header:
                continue
            if header in _UNIT_HEADERS:
                columns[column_index] = "unit"
            elif header in _RATE_HEADERS:
                columns[column_index] = rate_field
            elif header in _HEADER_ALIASES:
                columns[column_index] = _HEADER_ALIASES[header]
        has_item = FIELD_ITEM_IDENTIFIER in columns.values()
        has_value = any(
            field in _QUANTITY_FIELDS or field in _MONEY_FIELDS for field in columns.values()
        )
        if has_item and has_value:
            return row_index, columns
    return None, {}


def _read_rows(
    grid: list[list[object]], header_row: int, columns: dict[int, str], sheet_name: str
) -> list[SheetRow]:
    item_column = next(
        (index for index, field in columns.items() if field == FIELD_ITEM_IDENTIFIER), None
    )
    if item_column is None:  # pragma: no cover - _find_header guarantees one
        return []
    unit_column = next((index for index, field in columns.items() if field == "unit"), None)

    rows: list[SheetRow] = []
    for offset, raw in enumerate(grid[header_row + 1 :], start=header_row + 2):
        if item_column > len(raw):
            continue
        identifier = raw[item_column - 1]
        if identifier is None:
            continue
        item = str(identifier).strip()
        if not _ITEM_IDENTIFIER.match(item):
            # A total, a heading, or a note. Not an item.
            continue

        unit = None
        if unit_column is not None and unit_column <= len(raw):
            raw_unit = raw[unit_column - 1]
            unit = str(raw_unit).strip() if raw_unit is not None else None

        facts: list[SheetFact] = [
            SheetFact(
                fact_type=FIELD_ITEM_IDENTIFIER,
                kind=FactKind.IDENTIFIER,
                literal=item,
                value=None,
                unit=None,
                currency=None,
                cell=SheetCell(sheet=sheet_name, row=offset, column=item_column),
            )
        ]
        for column_index, field in sorted(columns.items()):
            if field in {FIELD_ITEM_IDENTIFIER, "unit"} or column_index > len(raw):
                continue
            cell_value = raw[column_index - 1]
            if cell_value is None:
                continue
            cell = SheetCell(sheet=sheet_name, row=offset, column=column_index)
            if field == FIELD_ITEM_DESCRIPTION:
                facts.append(
                    SheetFact(
                        fact_type=field,
                        kind=FactKind.TEXT,
                        literal=str(cell_value).strip(),
                        value=None,
                        unit=None,
                        currency=None,
                        cell=cell,
                    )
                )
                continue
            number = _to_decimal(cell_value)
            if number is None:
                continue
            is_money = field in _MONEY_FIELDS
            facts.append(
                SheetFact(
                    fact_type=field,
                    kind=FactKind.MONEY if is_money else FactKind.QUANTITY,
                    literal=str(cell_value).strip(),
                    value=number,
                    unit=None if is_money else unit,
                    currency=CURRENCY_INR if is_money else None,
                    cell=cell,
                )
            )

        if len(facts) > 1:
            rows.append(SheetRow(item_identifier=item, row_number=offset, facts=tuple(facts)))
    return rows
