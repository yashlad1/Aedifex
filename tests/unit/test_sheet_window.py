"""The spreadsheet window: a readable region of a workbook, bounded on every axis.

Justified on two grounds. It is a **resource limit** on untrusted input reached from an HTTP
request — a workbook with a million rows must not become a million-row response — and it is the
surface that made spreadsheet evidence checkable at all, so what it shows has to be what the
extractor read rather than a second interpretation of it.

Not tested here: whether the grid *looks* right. That is the viewer's problem, and it was checked by
opening it.
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import Workbook

from aedifex.errors import ExtractionError
from aedifex.extraction.spreadsheet import (
    MAX_WINDOW_COLUMNS,
    MAX_WINDOW_ROWS,
    cell_reference,
    read_sheet_window,
)


def _workbook(rows: int = 40, columns: int = 8, sheets: tuple[str, ...] = ("BOQ",)) -> bytes:
    book = Workbook()
    # `active` is optional to a type checker; creating sheets by name and dropping the default one
    # avoids asserting on something openpyxl always provides.
    del book["Sheet"]
    for name in sheets:
        sheet = book.create_sheet(name)
        for row in range(1, rows + 1):
            for column in range(1, columns + 1):
                sheet.cell(row=row, column=column, value=f"{name}-{row}-{column}")
    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


class TestWindowing:
    def test_it_centres_on_the_cited_row(self) -> None:
        window = read_sheet_window(_workbook(), sheet="BOQ", row=20, radius=3)

        assert [entry.number for entry in window.rows] == [17, 18, 19, 20, 21, 22, 23]
        assert window.truncated, "a sheet with rows outside the window must say so"

    def test_a_row_near_the_top_does_not_produce_row_zero(self) -> None:
        window = read_sheet_window(_workbook(), row=2, radius=5)

        assert window.first_row == 1
        assert window.rows[0].number == 1

    def test_the_window_is_bounded_however_large_the_request(self) -> None:
        """The caller is an HTTP request and the sheet is untrusted input."""
        window = read_sheet_window(_workbook(rows=4_000), row=2_000, radius=10_000)

        assert len(window.rows) <= MAX_WINDOW_ROWS
        assert all(len(entry.cells) <= MAX_WINDOW_COLUMNS for entry in window.rows)

    def test_columns_are_bounded_too(self) -> None:
        window = read_sheet_window(_workbook(columns=40), row=1, radius=0, columns=99)

        assert len(window.rows[0].cells) == MAX_WINDOW_COLUMNS


class TestFidelity:
    def test_a_cell_carries_the_reference_the_extractor_would_have_written(self) -> None:
        """The window and the facts must spell a location identically, or "where is this?" has two
        answers."""
        window = read_sheet_window(_workbook(), sheet="BOQ", row=6, radius=0)

        cell = window.rows[0].cells[5]
        assert cell.reference == cell_reference("BOQ", 6, 6)
        assert cell.letter == "F"

    def test_values_are_the_text_the_library_read(self) -> None:
        book = Workbook()
        sheet = book.create_sheet("BOQ")
        del book["Sheet"]
        sheet["A1"] = 853981318.41
        sheet["B1"] = None
        buffer = BytesIO()
        book.save(buffer)

        window = read_sheet_window(buffer.getvalue(), row=1, radius=0)

        # Not locale-formatted: showing "85,39,81,318" where the workbook holds this would present
        # our formatting as the document's content.
        assert window.rows[0].cells[0].value == "853981318.41"
        assert window.rows[0].cells[1].value == "", "an empty cell is empty, not 'None'"

    def test_every_sheet_is_listed_so_a_reader_can_move_between_them(self) -> None:
        window = read_sheet_window(_workbook(sheets=("BOQ", "Measurement")), row=1)

        assert window.sheets == ("BOQ", "Measurement")
        assert window.sheet == "BOQ", "the first sheet is the default, as it is for extraction"

    def test_a_named_sheet_is_honoured(self) -> None:
        window = read_sheet_window(
            _workbook(sheets=("BOQ", "Measurement")), sheet="Measurement", row=2, radius=0
        )

        assert window.sheet == "Measurement"
        assert window.rows[0].cells[0].value == "Measurement-2-1"


class TestRefusals:
    def test_an_unknown_sheet_is_refused_by_name(self) -> None:
        with pytest.raises(ExtractionError, match="no sheet named"):
            read_sheet_window(_workbook(), sheet="Nope")

    def test_bytes_that_are_not_a_workbook_are_refused(self) -> None:
        with pytest.raises(ExtractionError, match="could not be opened"):
            read_sheet_window(b"%PDF-1.7 this is not a spreadsheet")
